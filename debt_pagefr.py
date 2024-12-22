from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QTableWidget,
    QTableWidgetItem,
    QSizePolicy,
    QHeaderView,
    QMessageBox,
    QDialog,
    QPushButton,
    QHBoxLayout,
    QLabel,
)
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtGui import QIcon
import cx_Oracle
from AddDebtDialogfr import AddDebtDialog
from datetime import datetime
from EditDebtDialog import EditDebtDialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from PyQt5.QtCore import QSize
from datetime import datetime
from customer_detail_page import CustomerDebtWindow
from customer import Customer
from debt import Debt


def connect_to_db():
    dsn = cx_Oracle.makedsn("localhost", "1521", service_name="MANAGEMENT3")
    return cx_Oracle.connect(user="admin", password="2024", dsn=dsn)


class DebtPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.setWindowTitle("Gestion des dettes")
        # self.setLayoutDirection(Qt.RightToLeft)
        self.setStyleSheet(self.load_stylesheet())
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Table Widget
        self.table = QTableWidget()
        self.table.setColumnCount(6)  # Augmenté le nombre de colonnes pour les boutons
        self.table.setHorizontalHeaderLabels(
            [
                "Nom",
                "Montant total",
                "Date de la dette",
                "Dettes payées",
                "Dettes actuelles",
                "Actions",
            ]
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        layout.addWidget(self.table)

        # Ajouter le bouton "Ajouter"
        add_button = QPushButton()
        add_button.setIcon(QIcon("add_icon.png"))  # Utilisez votre fichier d'icône ici
        add_button.setText(
            "Ajouter"
        )  # Optionnel : vous pouvez ajouter du texte sous l'icône
        add_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        add_button.clicked.connect(
            lambda: self.add_customer()
        )  # Appeler add_debt avec une valeur de ligne fictive
        layout.addWidget(
            add_button, alignment=Qt.AlignBottom | Qt.AlignRight
        )  # Aligner le bouton en bas à droite

        self.setLayout(layout)

        self.load_debt_data()
        self.table.cellChanged.connect(self.on_cell_changed)

        # Connect double-click signal to show debt details
        self.table.cellDoubleClicked.connect(self.show_debt_details)

        # Centrer le widget sur le parent si présent
        if self.parent_window:
            self.center_on_parent(self.parent_window)

    def load_stylesheet(self):
        return """
            QWidget {
            }
            QPushButton {
                background-color: #007BFF;
                color: white;
                padding: 10px;
                border: none;
                border-radius: 15px;
                font-size: 12px;
                font-weight: bold;
                min-width: 10px;
                min-height: 10px;
            }
            QPushButton:hover {
                background-color: #0056b3;
                
            }
            QPushButton:pressed {
                background-color: #004085;
            }
            QTableWidget {
                padding: 1px;
                border: 1px solid #ddd;
                background-color: #ffffff;
                alternate-background-color: #f5f5f5;
                min-width: 10px;
                min-height: 10px;
            }
            QTableWidget QHeaderView::section {
                background-color: #007BFF;
                color: white;
                padding: 8px;
                font-size: 14px;
                border: 1px solid #007BFF;
                font-weight: bold;
                text-align: center;
            }
            QTableWidget::item {
                padding: 1px;
                border: 1px solid #ddd;
            }
            QTableWidget::item:selected {
                background-color: #007BFF;
                color: black;
            }
            QTableWidget::item:hover {
                background-color: #e9ecef;
            }
            QTableWidget::verticalHeader {
                border: 1px solid #ddd;
                visible: false;
            }
        """

    def center_on_parent(self, parent):
        parent_rect = parent.rect()
        widget_rect = self.rect()
        center_point = parent_rect.center()
        widget_position = QPoint(
            center_point.x() - widget_rect.width() // 2,
            center_point.y() - widget_rect.height() // 2,
        )
        self.move(widget_position)

    def load_debt_data(self):
        print(
            "Chargement des données de dettes depuis la base de données..."
        )  # Impression de débogage
        with connect_to_db() as connection:
            cursor = connection.cursor()
            cursor.execute(
                """ 
                SELECT MAX(d.ID) AS ID,d.CUSTOMER_ID,c.NAME,
               MAX(d.AMOUNT) KEEP (DENSE_RANK LAST ORDER BY d.ID DESC) AS AMOUNT,
               MAX(d.DEBT_DATE) KEEP (DENSE_RANK LAST ORDER BY d.ID DESC) AS DEBT_DATE,
              MAX(d.PAID_DEBT) KEEP (DENSE_RANK LAST ORDER BY d.ID DESC) AS PAID_DEBT,
              MAX(d.CURRENT_DEBT) KEEP (DENSE_RANK LAST ORDER BY d.ID DESC) AS CURRENT_DEBT
              FROM DEBTS d JOIN CUSTOMER c ON d.CUSTOMER_ID = c.ID GROUP BY d.CUSTOMER_ID, c.NAME ORDER BY ID DESC """
            )
            rows = cursor.fetchall()
            print(f"Récupéré {len(rows)} lignes de la base de données.")
            self.table.blockSignals(
                True
            )  # Bloquer les signaux avant de faire des changements
            self.table.setRowCount(len(rows))

            # Dictionnaires pour stocker les valeurs précédentes
            self.previous_paid = {row: 0.0 for row in range(len(rows))}
            self.debt_ids = {row: 0 for row in range(len(rows))}
            self.customer_ids = {row: 0 for row in range(len(rows))}

            for row_idx, row in enumerate(rows):
                for col_idx, item in enumerate(
                    row[2:]
                ):  # Commencer à partir du deuxième élément
                    table_item = QTableWidgetItem(
                        f"{item:.2f}" if isinstance(item, (int, float)) else str(item)
                    )
                    table_item.setTextAlignment(Qt.AlignCenter)
                    if col_idx in [
                        0,
                        1,
                        2,
                        4,
                    ]:  # Rendre certaines cellules non éditables
                        table_item.setFlags(table_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(row_idx, col_idx, table_item)

                # Ajouter des boutons pour chaque ligne
                self.add_button_to_row(
                    row_idx, row[0]
                )  # Passer l'index de la ligne et l'ID de la dette
                self.table.setRowHeight(row_idx, 50)

                self.previous_paid[row_idx] = float(row[5]) if row[5] else 0.0
                self.debt_ids[row_idx] = row[0]
                self.customer_ids[row_idx] = row[1]
                print(
                    f"Ligne {row_idx} -> ID de la dette : {row[0]}"
                )  # Impression de débogage
                print(
                    f"Ligne {row_idx} -> ID de la dette : {row[1]}"
                )  # Impression de débogage
                print(f"Ligne {row_idx} -> Dette payée : {row[4]}")

            self.table.blockSignals(
                False
            )  # Réactiver les signaux après que les changements sont terminés

    def add_button_to_row(self, row, debt_id):
        # Créer une mise en page horizontale pour les boutons
        button_layout = QHBoxLayout()

        # Ajouter le bouton "Supprimer" avec icône
        delete_button = QPushButton()
        delete_button.setIcon(
            QIcon("delete_icon.png")
        )  # Utilisez votre fichier d'icône ici
        delete_button.setSizePolicy(
            QSizePolicy.Fixed, QSizePolicy.Fixed
        )  # Fixer la taille du bouton
        delete_button.setFixedSize(30, 30)  # Définir une taille fixe pour le bouton
        delete_button.setIconSize(
            QSize(24, 24)
        )  # Ajuster la taille de l'icône si nécessaire
        delete_button.setText("")  # Effacer le texte pour ne montrer que l'icône
        delete_button.clicked.connect(lambda: self.delete_debt(row))
        button_layout.addWidget(delete_button)

        # Créer un QWidget pour contenir la mise en page
        button_widget = QWidget()
        button_widget.setLayout(button_layout)
        # Ajouter la mise en page à la dernière colonne du tableau
        self.table.setCellWidget(row, 5, button_widget)

    def on_cell_changed(self, row, column):
        if column == 3:  # Colonne pour la dette actuelle
            print(f"Cellule changée à la ligne {row}, colonne {column}")
            self.handle_paid_debt_change(row, column)

    def handle_paid_debt_change(self, row, column):
        input_item = self.table.item(row, column)
        paid_value = input_item.text() if input_item else ""
        print(
            f"Gérer le changement de dette payée : {paid_value}"
        )  # Impression de débogage

        if paid_value == "":  # Si la cellule est vide
            self.prompt_for_previous_value(row, column)
            return

        try:
            paid_value = float(paid_value)
            print(f"Valeur payée convertie en float : {paid_value}")
        except ValueError:
            self.table.blockSignals(True)
            self.table.setItem(row, column, QTableWidgetItem("0.00"))
            self.table.blockSignals(False)
            return

        # Mettre à jour le solde
        self.calculate_current_debt(row, column, paid_value)

    def prompt_for_previous_value(self, row, column):
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Recalcul")
        msg_box.setText("Aucune valeur n'a été saisie. Utiliser la valeur précédente ?")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        result = msg_box.exec_()

        if result == QMessageBox.Yes:
            previous_value = self.previous_paid.get(row, 0.0)
            self.table.blockSignals(True)
            self.table.setItem(row, column, QTableWidgetItem(f"{previous_value:.2f}"))
            self.table.blockSignals(False)
            self.calculate_current_debt(row, column, previous_value)

    def calculate_current_debt(self, row, column, paid_value):
        total_amount = float(self.table.item(row, 1).text())  # Montant total
        current_debt_item = self.table.item(row, 4)  # Dette actuelle
        current_debt = total_amount - paid_value
        if current_debt < 0:
            current_debt = 0.0
        print(f"Calcul de la dette actuelle pour la ligne {row}: {current_debt}")
        self.table.blockSignals(True)
        current_debt_item.setText(f"{current_debt:.2f}")
        self.table.blockSignals(False)
        # Update the database
        self.update_debt_in_database(row, column, paid_value, current_debt)

    def update_debt_in_database(self, row, column, paid_value, current_debt):
        debt_id = self.debt_ids[row]
        with connect_to_db() as connection:
            cursor = connection.cursor()
            cursor.execute(
                """SELECT CUSTOMER_ID,AMOUNT, DEBT_DATE, CUSTOMER_ID, CREATED_BY,
                      CREATED_AT, UPDATED_BY, UPDATED_AT FROM DEBTS WHERE ID = :debt_id""",
                {"debt_id": debt_id},
            )
            values = cursor.fetchone()
            customer_id = values[0]
            amount = values[1]
            debt_date = values[2]

        with connect_to_db() as connection:
            cursor = connection.cursor()
            cursor.execute(
                """INSERT INTO DEBTS (AMOUNT, DEBT_DATE, PAID_DEBT, CURRENT_DEBT, CUSTOMER_ID, CREATED_BY,
                      CREATED_AT, UPDATED_BY, UPDATED_AT) VALUES (:amount, :debt_date, :paid_debt, :current_debt, :customer_id,
                      :created_by, :created_at, :updated_by, :updated_at)
                """,
                {
                    "amount": amount,
                    "debt_date": debt_date,  # Changer ':date' en ':debt_date' pour correspondre à la variable
                    "paid_debt": paid_value,
                    "current_debt": current_debt,  # Modifier ici pour faire référence à la variable amount pour CURRENT_DEBT
                    "customer_id": customer_id,
                    "created_by": 1,  # Remplacez par l'ID de l'utilisateur qui crée l'entrée
                    "created_at": debt_date,  # Correction de la clé pour correspondre à la colonne
                    "updated_by": 1,
                    "updated_at": debt_date,
                },
            )
            connection.commit()

    def delete_debt(self, row):
        debt_id = self.debt_ids.get(row)
        customer_id = self.customer_ids.get(row)
        if debt_id is None:
            return

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Confirmation")
        msg_box.setText("Êtes-vous sûr de vouloir supprimer cette dette ?")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        result = msg_box.exec_()

        if result == QMessageBox.Yes:
            with connect_to_db() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "DELETE FROM CUSTOMER WHERE ID = :customer_id",
                    {"customer_id": customer_id},
                )
                connection.commit()
                self.load_debt_data()
                print(f"Dette avec ID {customer_id} supprimée.")

    def add_customer(self):
        dialog = AddDebtDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            debtor_name, amount, debt_date, identity, telephone, date_naissance = (
                dialog.get_values()
            )
            debt_date = datetime.strptime(debt_date, "%Y-%m-%d")
            date_naissance = datetime.strptime(date_naissance, "%Y-%m-%d")

            with connect_to_db() as connection:
                cursor = connection.cursor()
                # Vérifier si un client avec les mêmes détails existe déjà
                cursor.execute(
                    """SELECT ID FROM CUSTOMER 
                   WHERE NAME = :name AND IDENTITE = :identity 
                   AND TELEPHONE = :telephone AND DATE_NAISS = :date_naissance""",
                    {
                        "name": debtor_name,
                        "identity": identity,
                        "telephone": telephone,
                        "date_naissance": date_naissance,
                    },
                )
                customer = cursor.fetchone()
                if customer is not None:
                    # Notifier l'utilisateur que le client existe déjà
                    msg_box = QMessageBox(self)
                    msg_box.setWindowTitle("Client Existant")
                    msg_box.setText(
                        "Un client avec les mêmes informations existe déjà."
                    )
                    msg_box.setStandardButtons(QMessageBox.Ok)
                    msg_box.exec_()
                    return

                # Si le client n'existe pas, créez un nouveau client
                cursor.execute(
                    """INSERT INTO CUSTOMER (NAME, IDENTITE, TELEPHONE, DATE_NAISS) 
                   VALUES (:name, :identity, :telephone, :date_naissance)""",
                    {
                        "name": debtor_name,
                        "identity": identity,
                        "telephone": telephone,
                        "date_naissance": date_naissance,
                    },
                )
                connection.commit()
                # Obtenir l'ID du nouveau client pour ajouter une dette
                cursor.execute(
                    "SELECT ID FROM CUSTOMER WHERE NAME = :name AND IDENTITE = :identity",
                    {"name": debtor_name, "identity": identity},
                )
                customer = cursor.fetchone()

                # Insérer la nouvelle dette avec l'ID du client
                customer_id = customer[0]
                print(
                    {
                        "amount": amount,
                        "debt_date": debt_date,
                        "paid_debt": 0,
                        "current_debt": amount,
                        "customer_id": customer_id,
                        "created_by": 1,
                        "created_at": debt_date,
                        "updated_by": 1,
                        "updated_at": debt_date,
                    }
                )

                cursor.execute(
                    """INSERT INTO DEBTS (AMOUNT, DEBT_DATE, PAID_DEBT, CURRENT_DEBT, CUSTOMER_ID, 
                                      CREATED_BY, CREATED_AT, UPDATED_BY, UPDATED_AT) 
                   VALUES (:amount, :debt_date, :paid_debt, :current_debt, :customer_id, 
                           :created_by, :created_at, :updated_by, :updated_at)""",
                    {
                        "amount": amount,
                        "debt_date": debt_date,
                        "paid_debt": 0,
                        "current_debt": amount,
                        "customer_id": customer_id,
                        "created_by": 1,  # Remplacez par l'ID de l'utilisateur
                        "created_at": debt_date,
                        "updated_by": 1,
                        "updated_at": debt_date,
                    },
                )
                connection.commit()
                self.load_debt_data()
                print(f"Nouvelle dette ajoutée pour {debtor_name}.")

    def add_debt(self):
        dialog = AddDebtDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            debtor_name, amount, debt_date, identity, telephone, date_naissance = (
                dialog.get_values()
            )
            debt_date = datetime.strptime(debt_date, "%Y-%m-%d")
            date_naissance = datetime.strptime(date_naissance, "%Y-%m-%d")
            with connect_to_db() as connection:
                cursor = connection.cursor()

                # Rechercher l'ID du client correspondant au nom du débiteur
                cursor.execute(
                    "SELECT ID FROM CUSTOMER WHERE NAME = :name", {"name": debtor_name}
                )
                customer = cursor.fetchone()
                print("###########customer:", customer)

                if customer is None:
                    # Créer un nouveau client si le client n'existe pas
                    cursor.execute(
                        "INSERT INTO CUSTOMER (NAME, IDENTITE, TELEPHONE, DATE_NAISS) VALUES (:name, :identity, :telephone, :date_naissance)",
                        {
                            "name": debtor_name,
                            "identity": identity,
                            "telephone": telephone,
                            "date_naissance": date_naissance,
                        },
                    )
                    connection.commit()  # Valider l'insertion

                    # Rechercher à nouveau l'ID du nouveau client
                    cursor.execute(
                        "SELECT ID FROM CUSTOMER WHERE NAME = :name",
                        {"name": debtor_name},
                    )
                    customer = cursor.fetchone()

                # Insérer la nouvelle dette avec l'ID du client
                customer_id = customer[0]  # Récupérer l'ID du client
                cursor.execute(
                    """INSERT INTO DEBTS (AMOUNT, DEBT_DATE, PAID_DEBT, CURRENT_DEBT, CUSTOMER_ID, CREATED_BY, CREATED_AT, UPDATED_BY, UPDATED_AT) VALUES (:amount, :debt_date, :paid_debt, :current_debt, :customer_id, :created_by, :created_at, :updated_by, :updated_at)
                """,
                    {
                        "amount": amount,
                        "debt_date": debt_date,  # Changer ':date' en ':debt_date' pour correspondre à la variable
                        "paid_debt": 0,
                        "current_debt": amount,  # Modifier ici pour faire référence à la variable amount pour CURRENT_DEBT
                        "customer_id": customer_id,
                        "created_by": 1,  # Remplacez par l'ID de l'utilisateur qui crée l'entrée
                        "created_at": debt_date,  # Correction de la clé pour correspondre à la colonne
                        "updated_by": 1,
                        "updated_at": debt_date,
                    },
                )
                connection.commit()
                self.load_debt_data()
                print(f"Nouvelle dette ajoutée pour {debtor_name}.")

    def export_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dettes"

        # En-têtes
        headers = [
            "Nom du débiteur",
            "Montant total",
            "Date de la dette",
            "Dettes payées",
            "Dettes actuelles",
        ]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

        # Ajouter les données de la table
        for row in range(self.table.rowCount()):
            for col in range(
                self.table.columnCount() - 1
            ):  # Ignorer la colonne des actions
                item = self.table.item(row, col)
                value = item.text() if item else ""
                sheet.cell(row=row + 2, column=col + 1, value=value)

        # Chemin du fichier à enregistrer
        file_path = os.path.expanduser("~/Desktop/dettes.xlsx")
        workbook.save(file_path)
        QMessageBox.information(
            self,
            "Exportation réussie",
            f"Données exportées avec succès vers {file_path}.",
        )
        print(f"Données exportées vers {file_path}.")

    def show_debt_details(self, row, column):
        if column == 0:  # Only respond to clicks on the name column
            debt_id = self.debt_ids.get(row)
            customer_id = self.customer_ids.get(row)
            if debt_id is not None:
                with connect_to_db() as connection:
                    cursor = connection.cursor()

                    cursor.execute(
                        """SELECT c.NAME,c.IDENTITE,c.TELEPHONE,c.DATE_NAISS,d.ID AS ID_debts,d.AMOUNT,d.DEBT_DATE,d.PAID_DEBT,
                    d.CURRENT_DEBT,c.ID AS CUSTOMER_ID,d.CREATED_BY,d.CREATED_AT,d.UPDATED_BY,
                   d.UPDATED_AT FROM DEBTS d  JOIN CUSTOMER c ON d.CUSTOMER_ID = c.ID WHERE d.CUSTOMER_ID = :id
                   """,
                        {"id": customer_id},
                    )

                    debt_details = cursor.fetchall()
                    print(debt_details)

                    if debt_details:
                        # Pass all the debts (multiple rows) to the dialog
                        self.show_details_dialog(debt_details)

    def show_details_dialog(self, debt_details_list):
        # Liste pour stocker les instances de Customer
        debts = []
        # Boucle sur chaque dette dans la liste des détails de dette
        for details in debt_details_list:
            # Création d'une instance de Customer pour chaque élément
            customer = Customer(
                name=details[0],  # Nom du client
                identite=details[1],  # Identité
                telephone=details[2],  # Téléphone
                date_naiss=details[3],  # Date de Naissance
            )
            debt = Debt(
                amount=details[5],  # Montant Total de la dette
                debt_date=details[6],  # Date de la Dette
                paid_debt=details[7],  # Dette Payée
                current_debt=details[8],  # Dette Actuelle
                customer_id=details[9],  # ID du client
                created_by=details[10],  # Créé par
                created_at=details[11],  # Date de Création
                updated_by=details[12],  # Mis à jour par
                updated_at=details[13],  # Date de Mise à jour
            )
            # Ajouter l'instance Customer à la liste
            debts.append(debt)
        customer_debt_window = CustomerDebtWindow(customer, debts)
        customer_debt_window.exec_()
