from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QSizePolicy,
    QHeaderView,
    QMessageBox,
    QDialog,
)
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtGui import QIcon
import cx_Oracle
from AddCurrencyDialog import AddCurrencyDialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os


"""def connect_to_db():
    dsn = cx_Oracle.makedsn("localhost", "1521", service_name="MANAGEMENT4")
    return cx_Oracle.connect(user="admin", password="2024", dsn=dsn)
"""


class CurrencyManagementPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent  # Reference to the MainWindow
        self.setWindowTitle("إدارة العملات")
        self.setLayoutDirection(Qt.RightToLeft)
        self.setStyleSheet(self.load_stylesheet())
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Button Layout
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(10)

        self.add_currency_button = QPushButton("إضافة عملة")
        self.add_currency_button.setObjectName("addButton")
        self.add_currency_button.clicked.connect(self.add_currency)

        self.delete_button = QPushButton("حذف")
        self.delete_button.setObjectName("deleteButton")
        self.delete_button.clicked.connect(self.delete_currency)

        self.ret_button = QPushButton("العودة")
        self.ret_button.setObjectName("returnButton")
        self.ret_button.clicked.connect(self.return_to_previous)

        button_layout.addStretch()
        button_layout.addWidget(self.add_currency_button)
        button_layout.addWidget(self.delete_button)
        button_layout.addWidget(self.ret_button)
        button_layout.addStretch()

        button_widget = QWidget()
        button_widget.setLayout(button_layout)

        layout.addWidget(button_widget)

        # Table Widget
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(
            ["اسم العملة", "الدخول", "الخروج", "الرصيد"]
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        layout.addWidget(self.table)
        self.setLayout(layout)

        #       self.load_currency_data()
        self.table.cellChanged.connect(self.on_cell_changed)

        # Center the widget on the parent if present
        if self.parent_window:
            self.center_on_parent(self.parent_window)

    def center_on_parent(self, parent):
        parent_rect = parent.rect()
        widget_rect = self.rect()
        center_point = parent_rect.center()
        widget_position = QPoint(
            center_point.x() - widget_rect.width() // 2,
            center_point.y() - widget_rect.height() // 2,
        )
        self.move(widget_position)

    def load_stylesheet(self):
        return """
            QWidget {
              
            }
            QPushButton {
                background-color: #007BFF;
                color: white;
                padding: 10px 15px;
                border: none;
                border-radius: 5px;
                font-size: 12px;
                font-weight: bold;
                min-width: 100px;
                min-height: 35px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
            QPushButton#addButton {
                background-color: #28a745;
            }
            QPushButton#deleteButton {
                background-color: #dc3545;
            }
            QPushButton#returnButton {
                background-color: #17a2b8;
            }
            QPushButton#addButton:hover {
                background-color: #218838;
            }
            QPushButton#deleteButton:hover {
                background-color: #c82333;
            }
            QPushButton#returnButton:hover {
                background-color: #138496;
            }
            QTableWidget {
                padding: 1px;
                border: 1px solid #ddd;
                background-color: #ffffff;
                alternate-background-color: #f5f5f5;
            }
            QTableWidget QHeaderView::section {
                background-color: #007BFF;
                color: white;
                padding: 8px;
                font-size: 14px;
                border: 1px solid #007BFF;
                font-weight: bold;
                text-align: center;
            }
            QTableWidget::item {
                padding: 1px;
                border: 1px solid #ddd;
            }
            QTableWidget::item:selected {
                background-color: #007BFF;
                color: black;
            }
            QTableWidget::item:hover {
                background-color: #e9ecef;
            }
            QTableWidget::verticalHeader {
                border: 1px solid #ddd;
                visible: false;
            }
        """

        # def load_currency_data(self):
        with connect_to_db() as connection:
            cursor = connection.cursor()
            cursor.execute(
                "SELECT ID, NAME, INPUT, OUTPUT, BALANCE FROM CURRENCIES ORDER BY ID DESC"
            )
            rows = cursor.fetchall()
            self.table.blockSignals(True)
            self.table.setRowCount(len(rows))
            # Dictionnaires pour stocker les valeurs précédentes
            self.previous_input = {row: 0.0 for row in range(self.table.rowCount())}
            self.previous_output = {row: 0.0 for row in range(self.table.rowCount())}
            self.currency_ids = {row: 0 for row in range(self.table.rowCount())}
            for row_idx, row in enumerate(rows):
                for col_idx, item in enumerate(row[1:]):
                    # table_item = QTableWidgetItem(str(item))
                    table_item = QTableWidgetItem(
                        f"{item:.2f}" if isinstance(item, (int, float)) else str(item)
                    )
                    table_item.setTextAlignment(Qt.AlignCenter)
                    if col_idx in [0, 3]:  # 0 pour 'name' et 3 pour 'balance'
                        table_item.setFlags(
                            table_item.flags() & ~Qt.ItemIsEditable
                        )  # Rendre la cellule non modifiable
                    self.table.setItem(row_idx, col_idx, table_item)
                # Store previous values for input and output columns
                self.previous_input[row_idx] = float(row[2]) if row[2] else 0.0
                self.previous_output[row_idx] = float(row[3]) if row[3] else 0.0
                self.currency_ids[row_idx] = row[0] if row[0] else 0
            self.table.blockSignals(False)

    def on_cell_changed(self, row, column):
        if column == 1:  # Column for inputs
            self.handle_input_output(row, column, True)
        elif column == 2:  # Column for outputs
            self.handle_input_output(row, column, False)

    def handle_input_output(self, row, column, is_input):
        input_item = self.table.item(row, column)
        input_value = input_item.text() if input_item else ""

        if input_value == "":  # If the cell is empty
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("إعادة الحساب")
            msg_box.setText("لم يتم إدخال أي قيمة. هل ترغب في استخدام القيمة السابقة؟")
            # Ajouter les boutons personnalisés avec les labels en arabe
            yes_button = msg_box.addButton("نعم", QMessageBox.AcceptRole)
            no_button = msg_box.addButton("لا", QMessageBox.RejectRole)

            # Définir le bouton par défaut
            msg_box.setDefaultButton(yes_button)
            reply = msg_box.exec_()  # Store the result in a variable

            if reply == QMessageBox.AcceptRole:
                input_value = (
                    self.previous_input[row] if is_input else self.previous_output[row]
                )
                # Update cell with previous value
                self.table.blockSignals(True)
                self.table.setItem(row, column, QTableWidgetItem(str(input_value)))
                self.table.blockSignals(False)
                # Recalculate balance with previous value
                self.calculate_balance(row, column, input_value, is_input)
            else:
                # Update cell with previous value without recalculating balance
                input_value = (
                    self.previous_input[row] if is_input else self.previous_output[row]
                )
                self.table.blockSignals(True)
                self.table.setItem(row, column, QTableWidgetItem(str(input_value)))
                self.table.blockSignals(False)
            return

        try:
            input_value = float(input_value)
        except ValueError:
            self.table.blockSignals(True)
            self.table.setItem(row, column, QTableWidgetItem("0.00"))
            self.table.blockSignals(False)
            return

        # Update balance
        self.calculate_balance(row, column, input_value, is_input)

    def calculate_balance(self, row, column, input_value, is_input):
        # Get the balance item
        balance_item = self.table.item(row, 3)
        if balance_item is None:
            # If the balance item is None, create a new item
            self.table.blockSignals(True)
            balance_item = QTableWidgetItem("0.00")
            self.table.setItem(row, 3, balance_item)
            self.table.blockSignals(False)
        # Get current balance value
        try:
            balance_value = float(balance_item.text())
        except ValueError:
            balance_value = 0.0

        if is_input:
            balance_value += input_value
            self.previous_input[row] = input_value
        else:
            balance_value -= input_value
            self.previous_output[row] = input_value

        self.table.blockSignals(True)
        balance_item.setText(f"{balance_value:.2f}")
        self.table.setItem(row, column, QTableWidgetItem(f"{input_value:.2f}"))
        self.table.blockSignals(False)

        # Update the database
        self.update_database(row, column, input_value, balance_value, is_input)

    def update_database(self, row, column, input_value, new_balance, is_input):
        currency_id = self.currency_ids[row]
        operation_name = self.table.item(row, 0).text()
        if is_input:
            update_query = "UPDATE CURRENCIES SET input = :value, balance = :balance WHERE name = :name and ID =:c_id"
        else:
            update_query = "UPDATE CURRENCIES SET output = :value, balance = :balance WHERE name = :name and ID =:c_id"

        """with connect_to_db() as connection:
            cursor = connection.cursor()
            cursor.execute(
                update_query,
                value=input_value,
                balance=new_balance,
                name=operation_name,
                c_id=currency_id,
            )
            connection.commit()"""

    def add_currency(self):
        dialog = AddCurrencyDialog()
        if dialog.exec_() == QDialog.Accepted:
            name, input_value, output_value, balance = dialog.get_values()
            print(name, input_value, output_value, balance)
            """with connect_to_db() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    "INSERT INTO currencies (name, input, output, balance) VALUES (:name, :input, :output, :balance)",
                    {
                        "name": name,
                        "input": float(input_value),
                        "output": float(output_value),
                        "balance": float(balance),
                    },
                )
                connection.commit()
            self.load_currency_data()"""

    def delete_currency(self):
        selected_row = sorted(
            set(item.row() for item in self.table.selectedItems() if item)
        )
        if not selected_row:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("تنبيه")
            msg_box.setText("يرجى تحديد صف واحد على الأقل للحذف.")
            confirm_button = msg_box.addButton("أوكي", QMessageBox.AcceptRole)
            msg_box.exec_()
            return
        currency_name = self.table.item(selected_row[0], 0).text()
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("تأكيد الحذف")
        msg_box.setText(f"هل أنت متأكد أنك تريد حذف العملة {currency_name}؟")
        # Ajouter les boutons personnalisés avec les labels en arabe
        yes_button = msg_box.addButton("نعم", QMessageBox.AcceptRole)
        no_button = msg_box.addButton("لا", QMessageBox.RejectRole)
        # Définir le bouton par défaut
        msg_box.setDefaultButton(yes_button)
        reply = msg_box.exec_()  # Store the result in a variable

        if reply == QMessageBox.AcceptRole:
            """with connect_to_db() as connection:
            cursor = connection.cursor()
            for row in selected_row:
                print("row:", row)
                currency_id = self.currency_ids.get(row)
                if currency_id is None:
                    # self.load_ids_from_file()
                    # deposit_id = self.money_ids.get(row)
                    print("reload currencyid file")
                print("currency_id:", currency_id)
                cursor.execute(
                    "DELETE FROM CURRENCIES WHERE ID = :id_c", id_c=currency_id
                )
            connection.commit()"""
            self.load_currency_data()

    def return_to_previous(self):
        parent = self.parent_window
        self.export_to_excel()
        if parent and hasattr(parent, "stacked_widget"):
            cash_management_index = parent.stacked_widget.indexOf(
                parent.cash_management_page
            )
            if cash_management_index != -1:
                parent.stacked_widget.setCurrentIndex(cash_management_index)

    def export_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "قائمة العملات"

        # Set the direction of the sheet to RTL
        ws.sheet_view.rightToLeft = True

        # Ajouter les en-têtes
        headers = ["اسم العملة", "الدخول", "الخروج", "الرصيد"]

        # Définir les styles
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(
            start_color="007BFF", end_color="007BFF", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Écrire les en-têtes dans la feuille de calcul
        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border

        # Appliquer l'alignement à droite pour toutes les cellules de données
        rtl_alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=0, wrap_text=True
        )

        # Réorganiser les données en fonction de l'ordre inversé des colonnes
        for row_num in range(self.table.rowCount()):
            for col_num in range(self.table.columnCount()):
                item = self.table.item(row_num, col_num)
                if item is not None:
                    cell = ws.cell(
                        row=row_num + 2, column=col_num + 1, value=item.text()
                    )
                    cell.alignment = rtl_alignment
                    cell.border = thin_border

        # Ajuster les largeurs de colonnes
        for col_num in range(len(headers)):
            column_letter = chr(65 + col_num)
            ws.column_dimensions[column_letter].width = (
                20  # Ajuster la largeur selon les besoins
            )
        # Définir le répertoire de sauvegarde

        # Définir le répertoire de sauvegarde
        backup_dir = "backup"
        os.makedirs(backup_dir, exist_ok=True)  # Créer le répertoire s'il n'existe pas
        # Enregistrer le fichier Excel
        today = datetime.today().strftime("%Y-%m-%d")
        file_path = os.path.join(backup_dir, f"قائمة_العملات_{today}.xlsx")
        try:
            wb.save(file_path)
            # QMessageBox.information(self, "تصدير ناجح", f"تم تصدير البيانات بنجاح إلى {file_path}")
        except Exception as e:
            QMessageBox.critical(
                self, "خطأ", f"خطأ أثناء تسجيل البيانات\nرجاءا اغلق الملف {file_path}"
            )
