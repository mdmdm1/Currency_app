from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QTableWidget,
    QTableWidgetItem,
    QSizePolicy,
    QHeaderView,
    QMessageBox,
    QPushButton,
    QHBoxLayout,
    QLabel,
)
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtGui import QIcon
import cx_Oracle
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from AddDepositDialog1 import AddDepositDialog


def connect_to_db():
    dsn = cx_Oracle.makedsn("localhost", "1521", service_name="MANAGEMENT3")
    return cx_Oracle.connect(user="admin", password="2024", dsn=dsn)


class DepositPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.setWindowTitle("Gestion des dépôts")
        self.setStyleSheet(self.load_stylesheet())
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Table Widget
        self.table = QTableWidget()
        self.table.setColumnCount(6)  # Adjusted for actions column
        self.table.setHorizontalHeaderLabels(
            [
                "Nom",
                "Montant",
                "Date de dépôt",
                "Dépôt libéré",
                "Dette actuelle",
                "Actions",
            ]
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        layout.addWidget(self.table)

        # Add "Add Deposit" Button
        add_button = QPushButton("Ajouter un dépôt")
        add_button.setIcon(QIcon("add_icon.png"))  # Replace with your icon path
        add_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        add_button.clicked.connect(self.add_deposit)
        layout.addWidget(add_button, alignment=Qt.AlignBottom | Qt.AlignRight)

        self.setLayout(layout)
        self.load_deposit_data()

        # Optional: Center widget on parent if provided
        if self.parent_window:
            self.center_on_parent(self.parent_window)

    def load_stylesheet(self):
        return """
            QPushButton {
                background-color: #007BFF;
                color: white;
                padding: 10px;
                border: none;
                border-radius: 15px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
            QTableWidget {
                border: 1px solid #ddd;
                background-color: #ffffff;
                alternate-background-color: #f5f5f5;
            }
            QTableWidget QHeaderView::section {
                background-color: #007BFF;
                color: white;
                font-weight: bold;
                text-align: center;
            }
        """

    def center_on_parent(self, parent):
        parent_rect = parent.geometry()
        self.move(parent_rect.center() - self.rect().center())

    def load_deposit_data(self):
        print("Loading deposit data from database...")
        with connect_to_db() as connection:
            cursor = connection.cursor()
            cursor.execute(
                """
                SELECT c.NAME, d.AMOUNT, d.DEPOSIT_DATE, d.RELEASED_DEPOSIT, d.CURRENT_DEBT
                FROM DEPOSITS d
                JOIN CUSTOMER c ON d.CUSTOMER_ID = c.ID
            """
            )
            rows = cursor.fetchall()
            self.table.setRowCount(len(rows))

            for row_idx, row in enumerate(rows):
                for col_idx, data in enumerate(row):
                    item = QTableWidgetItem(str(data))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.table.setItem(row_idx, col_idx, item)

                # Add action buttons
                self.add_action_buttons(row_idx)

    def add_action_buttons(self, row):
        layout = QHBoxLayout()
        delete_button = QPushButton()
        delete_button.setIcon(QIcon("delete_icon.png"))
        delete_button.setFixedSize(30, 30)
        delete_button.clicked.connect(lambda: self.delete_deposit(row))
        layout.addWidget(delete_button)

        button_widget = QWidget()
        button_widget.setLayout(layout)
        self.table.setCellWidget(row, 5, button_widget)

    def add_deposit(self):
        dialog = AddDepositDialog()
        if dialog.exec_():
            person_name, amount, deposit_date = dialog.get_values()
            released_deposit = float(0.0)
            with connect_to_db() as connection:
                cursor = connection.cursor()
                cursor.execute(
                    """
                    INSERT INTO deposits (person_name, amount, deposit_date, released_deposit, current_debt)
                    VALUES (:person_name, :amount, TRUNC(TO_DATE(:deposit_date, 'DD-MON-YY')), :released_deposit, :current_debt)
                """,
                    {
                        "person_name": person_name,
                        "amount": amount,
                        "deposit_date": deposit_date,
                        "released_deposit": released_deposit,
                        "current_debt": amount,
                    },
                )
                connection.commit()

            self.load_deposit_data()  # Reload data after adding
            # self.save_ids_to_file()  # Save IDs to file

    def delete_deposit(self, row):
        # Logic for deleting deposit
        print(f"Delete Deposit for row {row} clicked!")
        QMessageBox.information(self, "Delete", f"Dépôt à la ligne {row} supprimé.")

    def export_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dépôts"

        headers = ["Nom", "Montant", "Date de dépôt", "Dépôt libéré", "Dette actuelle"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")

        for row_idx in range(self.table.rowCount()):
            for col_idx in range(self.table.columnCount() - 1):
                value = (
                    self.table.item(row_idx, col_idx).text()
                    if self.table.item(row_idx, col_idx)
                    else ""
                )
                sheet.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        file_path = os.path.expanduser("~/Desktop/deposits.xlsx")
        workbook.save(file_path)
        QMessageBox.information(
            self, "Export Successful", f"Exported data to {file_path}."
        )
        print(f"Data exported to {file_path}.")
